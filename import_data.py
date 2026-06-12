"""Import real stock data from Toko Harahap Excel files into Prediksi Stok.

Pipeline:
1. Clear existing sales data
2. Generate synthetic data up to Jun 8
3. Append real sales from Jun 9-11 Excel files
4. Set stock baselines from PAGI values
5. Retrain all Prophet models
"""
import json
import os
import sqlite3
import sys
from datetime import date, datetime, timedelta
from pathlib import Path

import openpyxl

PROJECT_ROOT = Path(__file__).parent
DB_PATH = PROJECT_ROOT / "data" / "prediksi.db"
PRODUCTS_PATH = PROJECT_ROOT / "products.json"
MODELS_DIR = PROJECT_ROOT / "data" / "models"
CACHE_DIR = PROJECT_ROOT / "data" / "import_cache"

# Excel files from Toko Harahap
EXCEL_FILES = {
    date(2026, 6, 9):  "data/doc_902ba7297760_STOK BARANG TOKO HARAHAP 9-6-26.xlsx",
    date(2026, 6, 10): "data/doc_729acc235a55_STOK BARANG TOKO HARAHAP 10-6-26.xlsx",
    date(2026, 6, 11): "data/doc_83c09a5882a5_STOK BARANG TOKO HARAHAP 11-6-26.xlsx",
}

# Product name mapping: Excel name -> system name
PRODUCT_MAP = {
    "GULA": "Gula",
    "MINYAK": "Minyak",
    "TEPUNG": "Tepung",
    "BERAS": "Beras",
    "AQUA": "Aqua",
    "ROTI HITAM MANIS": "Roti hitam manis",
    "GARAM": "Garam",
}


def parse_excel_date_label(path: str) -> date:
    """Extract date from the TGL cell in the Excel file."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    dt = ws.cell(2, 2).value  # Row 2, Col B = TGL
    if isinstance(dt, datetime):
        return dt.date()
    if isinstance(dt, date):
        return dt
    raise ValueError(f"Cannot parse date from {path}: {dt!r}")


def parse_excel_sales(path: str) -> dict[str, float]:
    """Parse Excel file and return {product_name: terjual} dict."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    sales = {}
    for row in ws.iter_rows(min_row=5, values_only=True):
        no, name, pagi, sore, terjual = row[0], row[1], row[2], row[3], row[4]
        if no is None or name is None:
            continue
        name = str(name).strip().upper()
        if name not in PRODUCT_MAP:
            print(f"  ⚠ Unknown product: {name}")
            continue
        # Parse TERJUAL value
        t = str(terjual).strip() if terjual else "0"
        # Handle formats like "0,7 TON/700 KG"
        # First try to extract the primary number
        import re
        match = re.search(r'^([\d,]+)', t.replace('.', ''))
        if match:
            qty_str = match.group(1).replace(',', '.')
            try:
                qty = float(qty_str)
            except ValueError:
                qty = 0.0
        else:
            qty = 0.0
        # Convert TON to KG for Beras
        sys_name = PRODUCT_MAP[name]
        if sys_name == 'Beras':
            qty *= 1000
        sales[sys_name] = qty
    return sales


def parse_excel_morning_stock(path: str) -> dict[str, float]:
    """Parse PAGI stock values for baseline."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    stocks = {}
    for row in ws.iter_rows(min_row=5, values_only=True):
        no, name, pagi, sore, terjual = row[0], row[1], row[2], row[3], row[4]
        if no is None or name is None:
            continue
        name = str(name).strip().upper()
        if name not in PRODUCT_MAP:
            continue
        # Parse PAGI value
        p = str(pagi).strip() if pagi else "0"
        import re
        match = re.search(r'^([\d,]+)', p.replace('.', ''))
        if match:
            qty_str = match.group(1).replace(',', '.')
            try:
                qty = float(qty_str)
            except ValueError:
                qty = 0.0
        else:
            qty = 0.0
        # Convert TON to KG for Beras
        sys_name = PRODUCT_MAP[name]
        if sys_name == 'Beras':
            qty *= 1000
        stocks[sys_name] = qty
    return stocks


def clear_data(conn: sqlite3.Connection) -> None:
    conn.execute("DELETE FROM sales_reports")
    conn.execute("DELETE FROM stock_snapshots")
    conn.execute("DELETE FROM outgoing_messages")
    conn.commit()
    print("  Cleared all existing data")


def generate_synthetic(conn: sqlite3.Connection, products: dict, up_to: date) -> int:
    """Generate synthetic sales data up to (but not including) up_to date."""
    from app.synthetic_data import generate_synthetic_data
    # We need days count from today's perspective up to up_to
    days = (date.today() - up_to).days + (up_to - date(2025, 12, 1)).days
    # Actually, generate for exactly up_to - some_start date
    # Let's calculate: from 90 days before up_to to up_to-1
    synthetic = generate_synthetic_data(products, days=90)
    count = 0
    for row in synthetic:
        row_date = datetime.fromisoformat(row["reported_at"]).date()
        if row_date < up_to:
            conn.execute(
                "INSERT INTO sales_reports (product_name, quantity, reported_at) VALUES (?, ?, ?)",
                (row["product_name"], row["quantity"], row["reported_at"]),
            )
            count += 1
    conn.commit()
    print(f"  Generated {count} synthetic sales rows up to {up_to}")
    return count


def insert_real_sales(conn: sqlite3.Connection, daily_sales: dict[date, dict[str, float]]) -> int:
    """Insert real sales data with timestamps."""
    count = 0
    for d, products_sales in sorted(daily_sales.items()):
        for product_name, qty in products_sales.items():
            if qty <= 0:
                continue
            # Create a timestamp at 18:00 on that day (after store closes)
            ts = datetime(d.year, d.month, d.day, 18, 0, 0).isoformat()
            conn.execute(
                "INSERT INTO sales_reports (product_name, quantity, reported_at) VALUES (?, ?, ?)",
                (product_name, qty, ts),
            )
            count += 1
    conn.commit()
    print(f"  Inserted {count} real sales rows")
    return count


def set_baseline_stocks(conn: sqlite3.Connection, morning_stock: dict[str, float], base_date: date) -> None:
    """Set initial stock confirmation from morning stock on Jun 9."""
    count = 0
    for product_name, qty in morning_stock.items():
        ts = datetime(base_date.year, base_date.month, base_date.day, 8, 0, 0).isoformat()
        conn.execute(
            "INSERT INTO stock_snapshots (product_name, quantity, snapshot_date, is_confirmation) VALUES (?, ?, ?, 1)",
            (product_name, qty, ts),
        )
        count += 1
    conn.commit()
    print(f"  Set baseline stock for {count} products")


def main() -> None:
    print("=" * 60)
    print("PREDIKSI STOK — DATA IMPORT")
    print("=" * 60)

    # 1. Load products
    with open(PRODUCTS_PATH) as f:
        products = json.load(f)
    print(f"\n📦 Products loaded: {len(products)}")

    # 2. Parse all Excel files
    daily_sales: dict[date, dict[str, float]] = {}
    morning_stock: dict | None = None
    first_date: date | None = None

    print(f"\n📄 Parsing Excel files...")
    for d, path in sorted(EXCEL_FILES.items()):
        if not os.path.exists(path):
            print(f"  ⚠ File not found: {path}")
            continue
        file_date = parse_excel_date_label(path)
        sales = parse_excel_sales(path)
        morning_stock_today = parse_excel_morning_stock(path)
        print(f"  {d}: {len(sales)} products with sales")
        daily_sales[d] = sales
        if first_date is None:
            first_date = d
            morning_stock = morning_stock_today

    if not daily_sales:
        print("❌ No Excel files processed!")
        return

    # 3. Connect to DB and import
    conn = sqlite3.connect(str(DB_PATH))
    try:
        print(f"\n💾 Importing to {DB_PATH}...")
        clear_data(conn)
        generate_synthetic(conn, products, first_date)
        insert_real_sales(conn, daily_sales)
        if morning_stock and first_date:
            set_baseline_stocks(conn, morning_stock, first_date)

        # Verify
        total = conn.execute("SELECT COUNT(*) FROM sales_reports").fetchone()[0]
        snap = conn.execute("SELECT COUNT(*) FROM stock_snapshots").fetchone()[0]
        rng = conn.execute("SELECT MIN(reported_at), MAX(reported_at) FROM sales_reports").fetchone()
        print(f"\n✅ Import complete!")
        print(f"   Total sales: {total} rows ({rng[0][:10]} → {rng[1][:10]})")
        print(f"   Stock snapshots: {snap}")
    finally:
        conn.close()

    # 4. Retrain models
    print(f"\n🧠 Retraining Prophet models...")
    from app.predictor import train_all_products
    train_all_products(str(DB_PATH), str(PRODUCTS_PATH), models_dir=str(MODELS_DIR))
    print(f"✅ All models retrained!")

    # 5. Print data summary
    print(f"\n📊 REAL SALES SUMMARY")
    print(f"   {'─'*50}")
    for d in sorted(daily_sales.keys()):
        print(f"   {d.strftime('%a %d/%m')}:")
        for pname, qty in sorted(daily_sales[d].items()):
            print(f"     {pname:20s} → {qty:8.0f}")
    print()


if __name__ == "__main__":
    main()
